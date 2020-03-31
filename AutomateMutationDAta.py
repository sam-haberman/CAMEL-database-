# Script that will take in a mutation csv from AleDb and convert it to the Camel Mutation format
# Starting by just doing one file
# Then move to go through all files in a folder

import csv
import re
from openpyxl import load_workbook
import pandas as pd
import shutil
import glob

path = "C:/Users/samue/Desktop/Thesis/ALEDB_conversion/Experiment_Files_to_convert"

def mutTranslate(csvfile):
    # Open each csvfile
    f = open(csvfile, 'r', encoding='UTF-8')
    csv_f = csv.reader(f)
    # Remove header row from csv file and put it in a list
    Isolates = []
    header = next(csv_f)
    index = 6
    while index < len(header):
        Isolates.append(header[index])
        index += 1
    # Use regular expressions to get the actual headers from the header file
    Groups = []
    for i in range(len(Isolates)):
        Groups.append(re.findall("[A][0-9]+[ ][F][0-9]+[ ][I][0-9]+[ ][R][0-9]+", Isolates[i]))
    # Add each mutation from csv file into a list
    data = []
    for row in csv_f:
        data.append(row)
    stors = []

    # # Loop to do Type of Mutation
    j = 0
    while j < len(data):
        # We are adding two extra columns to our csv file to match our template
        data[j].insert(4, "")
        data[j].insert(2, "")
        # If the mutation is type SNP we want to move the data so that the end position is equal to the starting
        # position + 1 as well as moving the data from looking like A->T to A in column 4 and T in column 5
        if data[j][3] == "SNP":
            temp = data[j][4]
            splitter = re.split("→", temp)
            data[j][4] = splitter[0]
            data[j][5] = splitter[1]
            data[j][2] = int(data[j][1].replace(',', "")) + 1
        # If the mutation is type Del we first check if it starts with a Δ, this represents how many bases are deleted
        # we remove the information from columns 4 and 5 and adjust the end position to be the starting position + the
        # number of bases deleted
        if data[j][3] == "DEL":
            if data[j][4][0] == "Δ":
                temp = re.split('Δ', data[j][4])
                newtemp = re.split(" ", temp[1])
                data[j][2] = int(data[j][1].replace(",", "")) + int(newtemp[0].replace(",", ""))
                data[j][4] = ""
                data[j][5] = ""
            # If the mutation information doesn't start with a Δ, then we collect the number of bases deleted and
            # add that value to our stat position to see what our end position is
            else:
                valuesplit = re.split("→", re.split("[(]|[)]", data[j][4])[2])
                vali = int(valuesplit[0]) - int(valuesplit[1])
                data[j][5] = ""
                data[j][4] = ""
                data[j][2] = int(data[j][1].replace(',', "")) + vali
        # If the mutation is type INS then if this starts with a + we find out how many bases are inserted and put them in
        # ALT and adjust the end position accordingly by adding one to the start position
        if data[j][3] == "INS":
            if data[j][4][0] == "+":
                data[j][5] = re.split("[+]", data[j][4])[1]
                data[j][4] = ""
                data[j][2] = int(data[j][1].replace(',', "")) + 1
            # If another format is used for INS then we do the same thing to see how many bases were inserted and add them
            # or the Insertion sequence identifier to ALT and adjust the end by adding one to the start position
            else:
                valuesplit = re.split("→", re.split("[(]|[)]", data[j][4])[2])
                vali = int(valuesplit[1]) - int(valuesplit[0])
                data[j][5] = re.split("[(]|[)]", data[j][4])[1] * vali
                data[j][4] = ""
                data[j][2] = int(data[j][1].replace(',', "")) + 1
        # If our mutation is type MOB then we parse the data to see how many bp were added and show what Insertion
        # sequence was added and how many bp it was, this is for the normal case of mobility
        if data[j][3] == "MOB":
            if data[j][4][0] == "I":
                temp = re.split(" [+]", data[j][4])
                data[j][4] = ""
                data[j][5] = temp[0]
                data[j][2] = int(data[j][1].replace(",", "")) + 1
            else:
                # For special case of mobility such as 	Δ2 bp :: IS186 (+) +9 bp we are treating as insertion of IS186
                data[j][3] = "INS"
                data[j][5] = re.split(" ", data[j][4])[3]
                data[j][4] = ""
                data[j][2] = int(data[j][1].replace(',', "")) + 1
        # If our mutation is type AMP than we update the name of Amplification to the number of times it is amplified
        # we also put in alt the length of the amplification
        if data[j][3] == "AMP":
            ampsplit = re.split(" ", data[j][4])
            data[j][3] = "AMP" + str(ampsplit[3])
            data[j][4] = ""
            data[j][5] = ampsplit[0]
            data[j][2] = int(data[j][1].replace(",", "")) + 1
            data[j][7] = ""
        # handling multiple substitutions by the number of bases changed
        if data[j][3] == "SUB":
            subsplit = re.split("→", data[j][4])
            data[j][2] = int(data[j][1].replace(',', "")) + int(re.split(" ",subsplit[0])[0])
            data[j][4] = ""
            data[j][5] = subsplit[1]
        # inversion mutations, inversion length is added to position start to signify length of the inversion
        if data[j][3] == "INV":
            invsplit = re.split(" ", data[j][4])
            data[j][2] = int(data[j][1].replace(',', "")) + int(re.split(" ",subsplit[0])[0])
            data[j][4] = ""
        # gene conversion mutations
        # if data[j][3] == "CON":
        #     #waiting for input

        j += 1
    # Loop to do multiple mutations per line
    j = 0
    # This loop checks each row of the Ale Db CSV data and sees how many experiment replicates this mutation occurred in
    # creating a new mutation for each replicate it occurred in
    while j < len(data):
        o = 8
        k = len(header) + 1
        while o < k:
            if data[j][o] != "":
                stors.append(
                    [data[j][0], data[j][1].replace(",", ""), data[j][2], data[j][3], data[j][4], data[j][5], data[j][6],
                     data[j][7].split(' ')[0], str(Groups[o - 8]).split(" ")[0][3:], str(Groups[o - 8]).split(" ")[2][1:],
                     str(Groups[o - 7]).split(" ")[1][1:], data[j][o], data[j][7]])
                o += 1
            else:
                o += 1
        j += 1
    # Final loop to clean up last columns
    # This is to clean up the template used in the Ale DB CSV to the template information we want in our data
    j = 0
    while j < len(stors):
        if stors[j][7] == "coding":
            stors[j][7] = ""
        elif stors[j][7] == "intergenic":
            stors[j][6] = "NA"
            stors[j][7] = ""
        elif (stors[j][3] == "DEL") and (stors[j][7] != ""):
            stors[j][7] = ""
        j += 1
    # This creates a data frame with the given columns and the fullly converted mutation data
    df = pd.DataFrame(stors,
                      columns=["CHROM", "START POS", "END POS", 'TYPE', 'REF', 'ALT', 'GEN', '∆AA', 'POP', 'CLON', 'TIME',
                               'FREQ', 'COM'])
    shutil.copy2("C:/Users/samue/Desktop/Thesis/Mutationlisttemplate_new.xlsx",
                 csvfile + '.xlsx')
    append_df_to_excel(csvfile + '.xlsx', df, index=False, header=False, startrow=5)
    wb = load_workbook(filename=csvfile + '.xlsx')
    ws = wb.worksheets[0]
    ws["K4"] = "Flask"
    wb.save(csvfile + '.xlsx')
    f.close()


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


# mutTranslate('C:/Users/samue/Desktop/Thesis/ALEDB_conversion/Experiment_Files_to_convert/42CTenaillonAra.csv')
# Goes through every file in the folder and runs our function to convert the CSV file into our template
# for file in glob.glob(path + '\\*'):
#     mutTranslate(file)
